from rest_framework import viewsets, permissions, filters
from rest_framework.views import APIView
from .models import MetroLine, Station, Position, Advertisement, AdvertisementArchive
from .serializers import (
    MetroLineSerializer, StationSerializer,
    PositionSerializer, AdvertisementSerializer, AdvertisementArchiveSerializer, CreateAdvertisementSerializer, ExportAdvertisementSerializer
)
from .pagination import CustomPagination
from rest_framework import status
from django.db import transaction
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import date, timedelta
from django.db.models import Q
from openpyxl import Workbook
from rest_framework.renderers import BaseRenderer
from drf_spectacular.utils import extend_schema, OpenApiParameter
from rest_framework.filters import SearchFilter
from django.utils import timezone



class XLSXRenderer(BaseRenderer):
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'xlsx'
    charset = None

    def render(self, data, media_type=None, renderer_context=None):
        return data



@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_me(request):
    return Response({
        "id": request.user.id,
        "username": request.user.username,
        "is_superuser": request.user.is_superuser,
        "email": request.user.email,
    })

class AuthenticatedCRUDPermission(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user and request.user.is_authenticated

class MetroLineViewSet(viewsets.ModelViewSet):
    queryset = MetroLine.objects.all()
    serializer_class = MetroLineSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    pagination_class = CustomPagination

class StationViewSet(viewsets.ModelViewSet):
    queryset = Station.objects.all()
    serializer_class = StationSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['line']
    # pagination_class = CustomPagination


class Stationimage(APIView):
    permission_classes = [IsAuthenticated]

    def put(self, request, pk):
        try:
            station = Station.objects.get(pk=pk)
        except Station.DoesNotExist:
            return Response({"error": "Bunday bekat topilmadi"}, status=404)

        image = request.FILES.get("schema_image")
        if not image:
            return Response({"error": "Rasm yuborilmadi"}, status=400)

        station.schema_image = image
        station.save()
        return Response({"message": "Rasm yangilandi", "id": station.id})
    

class PositionViewSet(viewsets.ModelViewSet):
    queryset = Position.objects.select_related('station').prefetch_related('advertisement').all()
    serializer_class = PositionSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    filterset_fields = ['station']
    search_fields = ['number']
    pagination_class = CustomPagination

    def perform_create(self, serializer):
        serializer.save(
            created_by=self.request.user,
            created_at=timezone.now()
        )

    def perform_destroy(self, instance):
        """
        Position o‘chirilsa -> unga biriktirilgan Advertisement bo‘lsa,
        avval arxivga ko‘chadi keyin o‘chadi, so‘ng Position o‘chiriladi.
        """
        ad = getattr(instance, 'advertisement', None)
        if ad:
            station = instance.station
            line = station.line if station else None

            AdvertisementArchive.objects.create(
                original_ad=ad,
                user=self.request.user,
                position=instance,
                line=line,
                station=station,
                Reklama_nomi=ad.Reklama_nomi,
                Qurilma_turi=ad.Qurilma_turi,
                Ijarachi=ad.Ijarachi,
                Shartnoma_raqami=ad.Shartnoma_raqami,
                Shartnoma_muddati_boshlanishi=ad.Shartnoma_muddati_boshlanishi,
                Shartnoma_tugashi=ad.Shartnoma_tugashi,
                O_lchov_birligi=ad.O_lchov_birligi,
                Qurilma_narxi=ad.Qurilma_narxi,
                Egallagan_maydon=ad.Egallagan_maydon,
                Shartnoma_summasi=ad.Shartnoma_summasi,
                Shartnoma_fayl=ad.Shartnoma_fayl,
                photo=ad.photo,
                contact_number=ad.contact_number,
            )
            ad.delete()

        instance.delete()

class AdvertisementViewSet(viewsets.ModelViewSet):
    queryset = Advertisement.objects.all()
    serializer_class = AdvertisementSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['Reklama_nomi', 'Shartnoma_raqami']
    ordering_fields = ['created_at', 'Qurilma_narxi']
    filterset_fields = ['position__station', 'position__station__line']
    # pagination_class = CustomPagination

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return CreateAdvertisementSerializer
        return AdvertisementSerializer

    def perform_create(self, serializer):
        serializer.save()

    def perform_update(self, serializer):
        old_instance = self.get_object()
        station = old_instance.position.station if old_instance.position else None
        line = station.line if station else None

        AdvertisementArchive.objects.create(
            original_ad=old_instance,
            user=self.request.user,
            position=old_instance.position,
            line=line,
            station=station,
            Reklama_nomi=old_instance.Reklama_nomi,
            Qurilma_turi=old_instance.Qurilma_turi,
            Ijarachi=old_instance.Ijarachi,
            Shartnoma_raqami=old_instance.Shartnoma_raqami,
            Shartnoma_muddati_boshlanishi=old_instance.Shartnoma_muddati_boshlanishi,
            Shartnoma_tugashi=old_instance.Shartnoma_tugashi,
            O_lchov_birligi=old_instance.O_lchov_birligi,
            Qurilma_narxi=old_instance.Qurilma_narxi,
            Egallagan_maydon=old_instance.Egallagan_maydon,
            Shartnoma_summasi=old_instance.Shartnoma_summasi,
            Shartnoma_fayl=old_instance.Shartnoma_fayl,
            photo=old_instance.photo,
            contact_number=old_instance.contact_number,
        )
        serializer.save(user=self.request.user)

    def perform_destroy(self, instance):
        """Delete tugmasi bosilganda reklama arxivga ko‘chadi va keyin o‘chadi"""
        station = instance.position.station if instance.position else None
        line = station.line if station else None

        AdvertisementArchive.objects.create(
            original_ad=instance,
            user=self.request.user,
            position=instance.position,
            line=line,
            station=station,
            Reklama_nomi=instance.Reklama_nomi,
            Qurilma_turi=instance.Qurilma_turi,
            Ijarachi=instance.Ijarachi,
            Shartnoma_raqami=instance.Shartnoma_raqami,
            Shartnoma_muddati_boshlanishi=instance.Shartnoma_muddati_boshlanishi,
            Shartnoma_tugashi=instance.Shartnoma_tugashi,
            O_lchov_birligi=instance.O_lchov_birligi,
            Qurilma_narxi=instance.Qurilma_narxi,
            Egallagan_maydon=instance.Egallagan_maydon,
            Shartnoma_summasi=instance.Shartnoma_summasi,
            Shartnoma_fayl=instance.Shartnoma_fayl,
            photo=instance.photo,
            contact_number=instance.contact_number,
        )
        instance.delete()

    @action(detail=True, methods=['get', 'post'], url_path='export')
    def export_advertisement(self, request, pk=None):
        source_ad = self.get_object()

        if request.method == 'GET':
            serializer = ExportAdvertisementSerializer()
            return Response(serializer.data)

        serializer = ExportAdvertisementSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        target_position = serializer.validated_data['position']

        with transaction.atomic():
            # 1. Target joydagi reklamani arxivga yuboramiz (agar mavjud bo‘lsa)
            target_ad = getattr(target_position, 'advertisement', None)
            if target_ad:
                target_station = target_ad.position.station
                target_line = target_station.line

                AdvertisementArchive.objects.create(
                    original_ad=target_ad,
                    user=self.request.user,
                    position=target_ad.position,
                    line=target_line,
                    station=target_station,
                    Reklama_nomi=target_ad.Reklama_nomi,
                    Qurilma_turi=target_ad.Qurilma_turi,
                    Ijarachi=target_ad.Ijarachi,
                    Shartnoma_raqami=target_ad.Shartnoma_raqami,
                    Shartnoma_muddati_boshlanishi=target_ad.Shartnoma_muddati_boshlanishi,
                    Shartnoma_tugashi=target_ad.Shartnoma_tugashi,
                    O_lchov_birligi=target_ad.O_lchov_birligi,
                    Qurilma_narxi=target_ad.Qurilma_narxi,
                    Egallagan_maydon=target_ad.Egallagan_maydon,
                    Shartnoma_summasi=target_ad.Shartnoma_summasi,
                    Shartnoma_fayl=target_ad.Shartnoma_fayl,
                    photo=target_ad.photo,
                    contact_number=target_ad.contact_number,
                )
                # 2. Target joydagi eski reklamani o‘chiramiz
                target_ad.delete()

            # 3. Source reklamani target pozitsiyaga nusxalaymiz
            Advertisement.objects.create(
                user=self.request.user,
                position=target_position,
                Reklama_nomi=source_ad.Reklama_nomi,
                Qurilma_turi=source_ad.Qurilma_turi,
                Ijarachi=source_ad.Ijarachi,
                Shartnoma_raqami=source_ad.Shartnoma_raqami,
                Shartnoma_muddati_boshlanishi=source_ad.Shartnoma_muddati_boshlanishi,
                Shartnoma_tugashi=source_ad.Shartnoma_tugashi,
                O_lchov_birligi=source_ad.O_lchov_birligi,
                Qurilma_narxi=source_ad.Qurilma_narxi,
                Egallagan_maydon=source_ad.Egallagan_maydon,
                Shartnoma_summasi=source_ad.Shartnoma_summasi,
                Shartnoma_fayl=source_ad.Shartnoma_fayl,
                photo=source_ad.photo,
                contact_number=source_ad.contact_number,
            )

            source_station = source_ad.position.station
            source_line = source_station.line

            AdvertisementArchive.objects.create(
                original_ad=source_ad,
                user=self.request.user,
                position=source_ad.position,
                line=source_line,
                station=source_station,
                Reklama_nomi=source_ad.Reklama_nomi,
                Qurilma_turi=source_ad.Qurilma_turi,
                Ijarachi=source_ad.Ijarachi,
                Shartnoma_raqami=source_ad.Shartnoma_raqami,
                Shartnoma_muddati_boshlanishi=source_ad.Shartnoma_muddati_boshlanishi,
                Shartnoma_tugashi=source_ad.Shartnoma_tugashi,
                O_lchov_birligi=source_ad.O_lchov_birligi,
                Qurilma_narxi=source_ad.Qurilma_narxi,
                Egallagan_maydon=source_ad.Egallagan_maydon,
                Shartnoma_summasi=source_ad.Shartnoma_summasi,
                Shartnoma_fayl=source_ad.Shartnoma_fayl,
                photo=source_ad.photo,
                contact_number=source_ad.contact_number,
            )

            # 5. Source joyni bo‘shatamiz
            source_ad.delete()

        return Response({'detail': 'Reklama muvaffaqiyatli ko‘chirildi, arxivlandi va source joy tozalandi.'}, status=200)
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Advertisements"

        headers = [
             "Reklama nomi", "Qurilma turi", "Ijarachi",
            "Shartnoma raqami", "Shartnoma boshlanishi", "Shartnoma tugashi",
            "O'lchov birligi", "Qurilma narxi", "Egallagan maydon", "Shartnoma summasi",
            "Position", "Station", "Line", "Contact number", "Created at", "Created by"
        ]
        ws.append(headers)

        for ad in queryset:
            row = [
                ad.Reklama_nomi,
                ad.Qurilma_turi,
                ad.Ijarachi,
                ad.Shartnoma_raqami,
                ad.Shartnoma_muddati_boshlanishi.strftime("%Y-%m-%d") if ad.Shartnoma_muddati_boshlanishi else "",
                ad.Shartnoma_tugashi.strftime("%Y-%m-%d") if ad.Shartnoma_tugashi else "",
                ad.O_lchov_birligi,
                float(ad.Qurilma_narxi) if ad.Qurilma_narxi else 0,
                ad.Egallagan_maydon,
                float(ad.Shartnoma_summasi) if ad.Shartnoma_summasi else 0,
                ad.position.number if ad.position else "",
                ad.position.station.name if ad.position and ad.position.station else "",
                ad.position.station.line.name if ad.position and ad.position.station and ad.position.station.line else "",
                ad.contact_number,
                ad.user.username if ad.user else "",
                ad.created_at.strftime("%Y-%m-%d %H:%M:%S") if ad.created_at else "",
            ]
            ws.append(row)

        # Column width optimize
        for i, column in enumerate(ws.columns, start=1):
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=advertisements.xlsx'
        wb.save(response)
        return response

class AdvertisementArchiveViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = AdvertisementArchive.objects.all().order_by('-created_at')
    serializer_class = AdvertisementArchiveSerializer
    permission_classes = [AuthenticatedCRUDPermission]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['Reklama_nomi','station__name']
    ordering_fields = ['created_at', 'Qurilma_narxi']
    filterset_fields = ['line','station', 'position']
    pagination_class = CustomPagination

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Advertisement Archives"

        headers = [
            "Original Ad ID", "Reklama nomi", "Qurilma turi", "Ijarachi",
            "Shartnoma raqami", "Shartnoma boshlanishi", "Shartnoma tugashi",
            "O'lchov birligi", "Qurilma narxi", "Egallagan maydon", "Shartnoma summasi",
            "Position", "Station", "Line", "User", "Created at", "Created by"
        ]
        ws.append(headers)

        for ad in queryset:
            row = [
                ad.original_ad.id if ad.original_ad else "",
                ad.Reklama_nomi,
                ad.Qurilma_turi,
                ad.Ijarachi,
                ad.Shartnoma_raqami,
                ad.Shartnoma_muddati_boshlanishi.strftime("%Y-%m-%d") if ad.Shartnoma_muddati_boshlanishi else "",
                ad.Shartnoma_tugashi.strftime("%Y-%m-%d") if ad.Shartnoma_tugashi else "",
                ad.O_lchov_birligi,
                float(ad.Qurilma_narxi) if ad.Qurilma_narxi else 0,
                ad.Egallagan_maydon,
                float(ad.Shartnoma_summasi) if ad.Shartnoma_summasi else 0,
                ad.position.number if ad.position else "",
                ad.station.name if ad.station else "",
                ad.line.name if ad.line else "",
                ad.user.username if ad.user else "",
                ad.created_at.strftime("%Y-%m-%d %H:%M:%S") if ad.created_at else "",
            ]
            ws.append(row)

        for i, column in enumerate(ws.columns, start=1):
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=advertisement_archives.xlsx'
        wb.save(response)
        return response
    

class ExpiredAdvertisementViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Advertisement.objects.all()
    serializer_class = AdvertisementSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter, DjangoFilterBackend]
    search_fields = ['Reklama_nomi','position__station__name']
    ordering_fields = ['Shartnoma_tugashi', 'Shartnoma_muddati_boshlanishi']
    pagination_class = CustomPagination

    def list(self, request, *args, **kwargs):
        today = date.today()
        seven_days_later = today + timedelta(days=7)

        # Asl querysetni search va filter bilan ishlatish
        queryset = self.filter_queryset(self.get_queryset())

        expired = queryset.filter(Shartnoma_tugashi__lt=today)
        expiring_soon = queryset.filter(
            Shartnoma_tugashi__range=(today, seven_days_later)
        )

        expired_data = self.get_serializer(expired, many=True).data
        expiring_soon_data = self.get_serializer(expiring_soon, many=True).data

        return Response({
            "counts": {
                "tugagan": expired.count(),
                "haftada_tugaydigan": expiring_soon.count(),
                "umumiy": queryset.count()
            },
            "results": {
                "tugagan": expired_data,
                "haftada_tugaydigan": expiring_soon_data
            }
        })

    @action(detail=False, methods=['get'], url_path='export-expired-excel')
    def export_expired_excel(self, request):
        today = date.today()
        expired = Advertisement.objects.filter(Shartnoma_tugashi__lt=today)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tugagan"

        headers = [
            "Reklama nomi", "Qurilma turi", "Ijarachi",
            "Shartnoma raqami", "Shartnoma boshlanishi", "Shartnoma tugashi",
            "Position", "Station", "Status", "Created by"
        ]
        ws.append(headers)

        for ad in expired:
            ws.append([
                ad.Reklama_nomi,
                ad.Qurilma_turi,
                ad.Ijarachi,
                ad.Shartnoma_raqami,
                ad.Shartnoma_muddati_boshlanishi.strftime("%Y-%m-%d") if ad.Shartnoma_muddati_boshlanishi else "",
                ad.Shartnoma_tugashi.strftime("%Y-%m-%d") if ad.Shartnoma_tugashi else "",
                ad.position.number if ad.position else "",
                ad.position.station.name if ad.position and ad.position.station else "",
                ad.user.username if ad.user else ""
                "tugagan"
            ])

        ws.append([])
        ws.append(["Umumiy tugagan soni:", expired.count()])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=expired_only.xlsx"
        wb.save(response)
        return response

    # ======================
    # Haftada tugaydiganlarni export qilish
    # ======================
    @action(detail=False, methods=['get'], url_path='export-week-excel')
    def export_week_excel(self, request):
        today = date.today()
        seven_days_later = today + timedelta(days=7)
        expiring_soon = Advertisement.objects.filter(
            Shartnoma_tugashi__range=(today, seven_days_later)
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Haftada_tugaydigan"

        headers = [
            "Reklama nomi", "Qurilma turi", "Ijarachi",
            "Shartnoma raqami", "Shartnoma boshlanishi", "Shartnoma tugashi",
            "Position", "Station", "Status" ,"Created by"
        ]
        ws.append(headers)

        for ad in expiring_soon:
            ws.append([
                ad.Reklama_nomi,
                ad.Qurilma_turi,
                ad.Ijarachi,
                ad.Shartnoma_raqami,
                ad.Shartnoma_muddati_boshlanishi.strftime("%Y-%m-%d") if ad.Shartnoma_muddati_boshlanishi else "",
                ad.Shartnoma_tugashi.strftime("%Y-%m-%d") if ad.Shartnoma_tugashi else "",
                ad.position.number if ad.position else "",
                ad.position.station.name if ad.position and ad.position.station else "",
                ad.user.username if ad.user else ""
                "haftada_tugaydigan"
            ])

        ws.append([])
        ws.append(["Umumiy haftada tugaydigan soni:", expiring_soon.count()])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=expiring_week.xlsx"
        wb.save(response)
        return response
    


class AllAdvertisementsViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Advertisement.objects.select_related("position__station__line", "user").all()
    serializer_class = AdvertisementSerializer
    permission_classes = [permissions.IsAuthenticated]   
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    
    # 3 ta qidiruv ustunlari
    search_fields = ["Ijarachi", "Shartnoma_raqami", "Qurilma_turi", "Reklama_nomi", "position__station__name"]
    filterset_fields = ["Ijarachi", "Shartnoma_raqami", "Qurilma_turi"]

    pagination_class = CustomPagination


    @extend_schema(
        parameters=[
            OpenApiParameter(name="search", description="Qidiruv uchun matn", required=False, type=str),
            OpenApiParameter(name="Ijarachi", description="Ijarachi bo‘yicha filter", required=False, type=str),
            OpenApiParameter(name="Shartnoma_raqami", description="Shartnoma raqami bo‘yicha filter", required=False, type=str),
        ],
        responses={200: 'Excel fayl'}
    )
    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        wb = Workbook()
        ws = wb.active
        ws.title = "Search Advertisements"

        headers = [
            "Reklama nomi", "Qurilma turi", "Ijarachi",
            "Shartnoma raqami", "Shartnoma boshlanishi", "Shartnoma tugashi",
            "O'lchov birligi", "Qurilma narxi", "Egallagan maydon", "Shartnoma summasi",
            "Position", "Station", "Line", "Contact number", "Created at", "Created by"
        ]
        ws.append(headers)

        for ad in queryset:
            ws.append([
                ad.Reklama_nomi,
                ad.Qurilma_turi,
                ad.Ijarachi,
                ad.Shartnoma_raqami,
                ad.Shartnoma_muddati_boshlanishi.strftime("%Y-%m-%d") if ad.Shartnoma_muddati_boshlanishi else "",
                ad.Shartnoma_tugashi.strftime("%Y-%m-%d") if ad.Shartnoma_tugashi else "",
                ad.O_lchov_birligi,
                float(ad.Qurilma_narxi) if ad.Qurilma_narxi else 0,
                ad.Egallagan_maydon,
                float(ad.Shartnoma_summasi) if ad.Shartnoma_summasi else 0,
                ad.position.number if ad.position else "",
                ad.position.station.name if ad.position and ad.position.station else "",
                ad.position.station.line.name if ad.position and ad.position.station and ad.position.station.line else "",
                ad.contact_number,
                ad.created_at.strftime("%Y-%m-%d %H:%M:%S") if ad.created_at else "",
                ad.user.username if ad.user else "",
            ])

        # ustun kengligi
        for i, column in enumerate(ws.columns, start=1):
            max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=search_advertisements.xlsx'
        wb.save(response)
        return response
